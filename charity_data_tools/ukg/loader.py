"""
Load and index a 360Giving UK Grantmaking (UKG) xlsx file.

The UKG dataset is published annually by 360Giving and covers all UK-based
grantmaking organisations.  Download the 'All Grantmakers' Excel file from
the 360Giving website and place it in your project's sources directory.

Expected file name pattern: ``UKGrantmaking-*-All-Grantmakers.xlsx``

Example::

    from pathlib import Path
    from charity_data_tools.ukg.loader import load_ukg_index, resolve_source_file

    source = resolve_source_file(sources_dir=Path("data/sources"))
    by_charity_num, by_orgid = load_ukg_index(source)

    record = by_charity_num.get("1234567")  # primary join key
    record = by_orgid.get("GB-CHC-1234567") # fallback join key
"""

from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl

from charity_data_tools.ukg.helpers import norm_charity_num


def resolve_source_file(
    arg: Optional[str] = None,
    sources_dir: Optional[Path] = None,
) -> Path:
    """Resolve the UKG source xlsx path.

    If *arg* is provided it is returned as a ``Path`` (error if the file
    does not exist).  Otherwise the most recently modified
    ``UKGrantmaking-*.xlsx`` in *sources_dir* is returned.

    Raises ``SystemExit`` with a clear message if the file cannot be found.
    """
    if arg:
        p = Path(arg).expanduser().resolve()
        if not p.exists():
            raise SystemExit("ERROR: --source-file not found: {}".format(p))
        return p
    if sources_dir is None:
        raise SystemExit(
            "ERROR: sources_dir must be provided when no explicit source file is given."
        )
    sd = Path(sources_dir)
    if not sd.exists():
        raise SystemExit(
            "ERROR: sources directory not found: {}\n"
            "  Create it and drop the latest UKGrantmaking-YYYY-All-Grantmakers.xlsx there,\n"
            "  or pass --source-file.".format(sd)
        )
    candidates = sorted(
        sd.glob("UKGrantmaking-*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise SystemExit(
            "ERROR: no UKGrantmaking-*.xlsx files found in {}\n"
            "  Download the latest from 360Giving and drop it there.".format(sd)
        )
    return candidates[0]


def load_ukg_index(
    source_path: Path,
) -> Tuple[Dict[str, Dict], Dict[str, Dict]]:
    """Read a UKG xlsx and return two lookup dicts pointing at the same row dicts.

    Returns ``(by_charity_num, by_orgid)`` where:

    - ``by_charity_num[normalised_charity_number]`` → row dict
    - ``by_orgid[org_id_string]`` → row dict

    Both dicts point at the same underlying objects — mutating a record via
    one index is reflected in the other.

    Raises ``SystemExit`` if the file lacks the expected sheet or columns.
    """
    print("Loading UKG source: {}".format(source_path))
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    if "Grantmaker" not in wb.sheetnames:
        raise SystemExit(
            "ERROR: source file has no 'Grantmaker' sheet. "
            "Found: {}".format(wb.sheetnames)
        )
    ws = wb["Grantmaker"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    h = {v: i for i, v in enumerate(header)}

    required = {"org_id", "Name", "REG: Charity Number"}
    missing = required - set(header)
    if missing:
        raise SystemExit(
            "ERROR: source file is missing expected columns: {}".format(missing)
        )

    by_charity_num: Dict[str, Dict] = {}
    by_orgid: Dict[str, Dict] = {}
    rows_total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[h["org_id"]] is None:
            continue
        rows_total += 1
        rec = {hdr: row[idx] for hdr, idx in h.items()}
        orgid = str(rec["org_id"]).strip()
        if orgid:
            by_orgid[orgid] = rec
        cn = norm_charity_num(rec.get("REG: Charity Number"))
        if cn:
            by_charity_num.setdefault(cn, rec)

    print("  UKG rows:          {}".format(rows_total))
    print("  Indexed by org_id: {}".format(len(by_orgid)))
    print("  Indexed by char#:  {}\n".format(len(by_charity_num)))
    return by_charity_num, by_orgid
